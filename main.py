import time
import json
import logging
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook


class SubFileHandler(FileSystemEventHandler):
    excel_path = None
    folder_path_tem = None
    folder_path = None
    folder_name_tem = None
    folder_name = None
    sheet_name = None

    def __init__(self):
        get_config(self)

    def on_created(self, event):
        work_on_excel(self, event)

    def on_deleted(self, event):
        work_on_excel(self, event)


def get_config(self, config_path="./config_tarkov.json"):
    try:

        # Read from config.
        f = open(config_path, mode='r', encoding='utf-8')
        config = json.loads(f.read())
        self.excel_path = config['excel_path']
        self.folder_path_tem = config['folder_path_tem']
        self.folder_path = config['folder_path']
        self.sheet_name = config['sheet_name']
        self.folder_name_tem = self.folder_path_tem.split('\\')[-1]
        self.folder_name = self.folder_path.split('\\')[-1]
        f.close()

        # Print last uploaded file.
        logging.info(
            'Last file uploaded was:\n' + '[' + config['last_file'] + "] " + config[
                'last_uploaded_file_time'])

    except IOError as msg:
        logging.error(msg)


def work_on_excel(self, event):
    try:
        wb = load_workbook(self.excel_path)
        try:
            ws = wb[self.sheet_name]

            if event.event_type == 'created':
                work_on_excel_by_event_created(self, ws, event)

            elif event.event_type == 'deleted':
                work_on_excel_by_event_deleted(self, ws, event)

            wb.save(self.excel_path)

        except IOError as msg:
            logging.error(msg)

    except IOError as msg:
        logging.error(msg)


def work_on_excel_by_event_created(self, ws, event, config_path="./config_tarkov.json"):
    current_row = ws.max_row + 1
    file_path = event.src_path.split('/')[-1]
    file_name = event.src_path.split('\\')[-1]
    try:
        # Modify these codes for private use.
        ws.cell(column=1, row=current_row, value="{0}".format(file_name)).hyperlink = file_path
        ws.cell(column=4, row=current_row, value="No")

        # Read from config.
        f = open(config_path, mode='r', encoding='utf-8')
        config = json.loads(f.read())
        f.close()

        # Rewrite config to update last file uploaded.
        f = open(config_path, mode='w', encoding='utf-8')
        config['last_file'] = file_name
        config['last_uploaded_file_time'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        f.write(json.dumps(config, indent=2, ensure_ascii=False))
        f.close()

        logging.info('Successfully write: ' + '[' + file_name + ']'
                     + ' From: ' + self.folder_name_tem + ' To ' + self.folder_name)

    except IOError as msg:
        logging.info(msg)


def work_on_excel_by_event_deleted(self, ws, event):
    try:
        file_name = event.src_path.split('\\')[-1]

        for row in range(2, ws.max_row + 1):
            if ws.cell(column=1, row=row).internal_value is not None:
                internal_value = ws.cell(column=1, row=row).internal_value
                if internal_value == file_name:
                    ws.delete_rows(idx=row)

                    # This is a potential openpyxl bug,
                    # that hyperlinks will stay at old places rather than going up after deleting rows,
                    # but they still hold the right place at some point.
                    # Have to concentrate their target again.
                    for row_sub in range(row, ws.max_row + 1):
                        ws.cell(column=1, row=row_sub).hyperlink = \
                            ws.cell(column=1, row=row_sub).hyperlink.target

        logging.info('Successfully delete: ' + '[' + file_name + ']' + ' From: ' + self.folder_name)

    except IOError as msg:
        logging.error(msg)


if __name__ == "__main__":
    print('Welcome to Python_folder_listener!')
    logging.basicConfig(level=logging.INFO)
    observer = Observer()
    event_handler = SubFileHandler()
    observer.schedule(event_handler, event_handler.folder_path, recursive=True)
    observer.start()
    try:
        while True:
            time.sleep(1)

    except KeyboardInterrupt:
        logging.info("Closed!")

    finally:
        observer.stop()
        observer.join()
